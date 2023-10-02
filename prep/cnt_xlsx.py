# # import os
# # import openpyxl
# # from PIL import Image

# # # 定义函数用于统计xlsx文件行数
# # def count_xlsx_lines(file_path):
# #     try:
# #         workbook = openpyxl.load_workbook(file_path, data_only=True)
# #         worksheet = workbook.active
# #         row_count = worksheet.max_row
# #         return row_count
# #     except Exception as e:
# #         print(f"Error counting lines in {file_path}: {e}")
# #         return 0

# # # 定义函数用于统计图片文件数量
# # def count_image_files(root_dir):
# #     image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp']
# #     image_count = 0
# #     for root, dirs, files in os.walk(root_dir):
# #         for file in files:
# #             _, ext = os.path.splitext(file)
# #             if ext.lower() in image_extensions:
# #                 image_count += 1
# #     return image_count

# # # 设置根目录
# # root_directory = '/mnt/task_runtime/data/ancient_char_data/单字/简帛_楚简字形库'  # 可以替换成您想要的根目录路径

# # # 初始化统计数据
# # total_xlsx_lines = 0
# # total_xlsx_files = 0  # 添加用于统计xlsx文件数量的变量
# # total_image_files = 0

# # # 遍历根目录下的文件和子目录
# # for root, dirs, files in os.walk(root_directory):
# #     for file in files:
# #         _, ext = os.path.splitext(file)
# #         if ext.lower() == '.xlsx':
# #             file_path = os.path.join(root, file)
# #             lines = count_xlsx_lines(file_path)
# #             total_xlsx_lines += lines
# #             total_xlsx_files += 1  # 增加xlsx文件数量
# #         elif ext.lower() in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']:
# #             total_image_files += 1

# # # 打印统计结果
# # print(f"总行数（xlsx文件）: {total_xlsx_lines}")
# # print(f"xlsx文件数量: {total_xlsx_files}")
# # print(f"图片文件数量: {total_image_files}")


# import os
# import openpyxl

# # 定义函数用于统计xlsx文件中不重复的字数
# def count_unique_chars_in_xlsx(file_path):
#     unique_chars = set()
#     try:
#         workbook = openpyxl.load_workbook(file_path, data_only=True)
#         for sheet in workbook.worksheets:
#             for row in sheet.iter_rows(values_only=True):
#                 for cell_value in row:
#                     if cell_value is not None and isinstance(cell_value, str):
#                         unique_chars.update(cell_value)
#         return len(unique_chars)
#     except Exception as e:
#         print(f"Error counting unique characters in {file_path}: {e}")
#         return 0

# # 设置根目录
# root_directory = '/mnt/task_runtime/data/ancient_char_data/单字/简帛_楚简字形库'  # 可以替换成您想要的根目录路径

# # 初始化统计数据
# total_unique_chars = 0

# # 遍历根目录下的文件
# for root, dirs, files in os.walk(root_directory):
#     for file in files:
#         _, ext = os.path.splitext(file)
#         if ext.lower() == '.xlsx':
#             file_path = os.path.join(root, file)
#             unique_chars_count = count_unique_chars_in_xlsx(file_path)
#             total_unique_chars += unique_chars_count

# # 打印统计结果
# print(f"所有xlsx文件中不重复的字数: {total_unique_chars}")




import os
import openpyxl
import re

# 用于匹配汉字的正则表达式
hanzi_pattern = re.compile(r'[\u4e00-\u9fa5]')

# 用于存储不重复的汉字
unique_hanzis = set()

# 设置根目录
root_directory = '/mnt/task_runtime/data/ancient_char_data/单字/简帛_楚简字形库'  # 可以替换成您要搜索的根目录路径

# 遍历根目录下的所有xlsx文件
for root, dirs, files in os.walk(root_directory):
    for file in files:
        _, ext = os.path.splitext(file)
        if ext.lower() == '.xlsx':
            file_path = os.path.join(root, file)
            try:
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    for row in sheet.iter_rows(values_only=True):
                        for cell_value in row:
                            if cell_value is not None and isinstance(cell_value, str):
                                hanzi_matches = hanzi_pattern.findall(cell_value)
                                unique_hanzis.update(hanzi_matches)
            except Exception as e:
                print(f"Error processing {file_path}: {e}")

# 统计不重复汉字数量
num_unique_hanzis = len(unique_hanzis)

# 打印结果
print(f"不重复汉字数量: {num_unique_hanzis}")
